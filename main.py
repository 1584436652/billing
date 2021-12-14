from openpyxl import load_workbook
from openpyxl import Workbook


def load_table(table_name, sheet_name):
    wb = load_workbook(table_name)
    ws = wb[sheet_name]
    # ws = wb.active
    rows = []
    sku_dict = dict()
    for row in ws.iter_rows():
        rows.append(row)
    for x in range(1, len(rows)):
        sku_name = str(rows[x][10].value)
        billing_weight = float(rows[x][13].value)
        if sku_name not in sku_dict:
            sku_dict[sku_name] = [billing_weight]
        else:
            sku_dict[sku_name].append(billing_weight)
    # print(sku_dict)
    return sku_dict


def billing_weight_sort(items):
    data = []
    for k, v in items.items():
        # 计费重排序
        so = sorted(v)
        # 最小计费重
        min_v = float(so[0])
        # 最大计费重
        max_v = float(so[-1])
        # 中位数
        size = len(so)
        # 衡量值
        measure = min_v * 1.3
        if size % 2 == 0:
            median = float(format((so[size // 2] + so[size // 2 - 1]) / 2, '.3f'))
        else:
            median = float(so[(size - 1) // 2])
        if max_v >= measure:
            abnormal = [k, so, min_v, max_v, median, measure, "异常"]
            data.append(abnormal)
        else:
            abnormal = [k, so, min_v, max_v, median, measure, None]
            data.append(abnormal)
    return data


def save(save_data):
    wb = Workbook()
    ws = wb.active
    ws.append(["SKU", "列表排序", "最小值", "最大值", "中间值", "衡量值", "abnormal"])
    print(f'总条数：{len(save_data)}')
    number = 2
    for save_value in save_data:
        ws[f'A{number}'] = save_value[0]
        ws[f'B{number}'] = str(save_value[1])
        ws[f'C{number}'] = save_value[2]
        ws[f'D{number}'] = save_value[3]
        ws[f'E{number}'] = save_value[4]
        ws[f'F{number}'] = save_value[5]
        ws[f'G{number}'] = save_value[6]
        wb.save('demo.xlsx')
        number += 1
    print("已保存")


if __name__ == '__main__':
    it = load_table("jifei.xlsx", "去除")
    sot = billing_weight_sort(it)
    save(sot)



