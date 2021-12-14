from openpyxl import load_workbook
from openpyxl import Workbook


class OperationForm:

    @staticmethod
    def load_table(table_path, sheet_name):
        try:
            return load_workbook(table_path)[sheet_name]
        except KeyError as key_e:
            print(f"{sheet_name} 不存在")
            raise key_e
        except FileNotFoundError as file_e:
            print(f"{table_path} 不存在")
            raise file_e

    def create_table(self):
        pass

    def transform_type(self):
        pass


class OperatingData(OperationForm):

    def load_data(self, *args):
        """
        获取表格里的数据
        :param args:
        :return:
        """
        load_ws = self.load_table(*args)
        li = list()
        for row in load_ws.iter_rows():
            li.append(row)
        for i in range(0, len(li)):
            detail = []
            for j in range(0, len(li[i])):
                detail.append(load_ws.cell(row=i + 1, column=j + 1).value)
            yield detail

    @staticmethod
    def conversion_dictionary(nested: list):
        """
        把嵌套的列表转成字典
        :param nested:
        :return:
        """
        detail = {}
        lens = len(nested[0])
        nested_len = len(nested)
        for m in range(0, lens):
            de = []
            for n in range(nested_len):
                de.append(nested[n][m])
                detail[de[0]] = de[1:]
        return detail

    @staticmethod
    def specified_value(va, *args):
        for v in args:
            yield {v: va[v]}


if __name__ == '__main__':
    data = []
    xl = OperatingData()
    sh = xl.load_data('test.xlsx', "Sheet1")
    for k in sh:
        data.append(k)
    di = xl.conversion_dictionary(data)
    for i in xl.specified_value(di, "运输方式", "国家"):
        print(i)

