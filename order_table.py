from openpyxl import load_workbook
from openpyxl import Workbook


"""
需要找出超出计重费对应的货运单号
"""


def load_table(table_name, sheet_name):
    wb = load_workbook(table_name)
    ws = wb[sheet_name]
    # ws = wb.active
    rows = []
    sku_dict = dict()
    for row in ws.iter_rows():
        rows.append(row)
    for x in range(1, len(rows)):
        order = str(rows[x][9].value)
        sku_name = str(rows[x][10].value)
        billing_weight = float(rows[x][13].value)
        if sku_name not in sku_dict:
            sku_dict[sku_name] = [{order: billing_weight}]
        else:
            sku_dict[sku_name].append({order: billing_weight})
    # print(sku_dict)
    return sku_dict


def billing_weight_sort(items):
    for k, v in items.items():
        weight = list()
        # 获取每个sku的计重费并且排序
        for m in v:
            for x, y in m.items():
                weight.append(y)
        weight_sort = sorted(weight)
        # 最小计费重
        min_v = float(weight_sort[0])
        # 最大计费v重
        max_v = float(weight_sort[-1])
        # 衡量值
        measure = float(format(min_v * 1.3, '.3f'))
        for o, p in enumerate(v):
            for a, b in p.items():
                if b >= measure:
                    print(k, a, b, min_v, max_v, measure)
                    yield [k, a, b, min_v, max_v, measure]


def exceed_table(exceed):
    wb = Workbook()
    ws = wb.active
    ws.append(["SKU", "货运单号", "计重费", "最小值", "最大值", "衡量值"])
    print(f'总条数：{len(exceed)}')
    number = 2
    for save_value in exceed:
        ws[f'A{number}'] = save_value[0]
        ws[f'B{number}'] = str(save_value[1])
        ws[f'C{number}'] = save_value[2]
        ws[f'D{number}'] = save_value[3]
        ws[f'E{number}'] = save_value[4]
        ws[f'F{number}'] = save_value[5]
        wb.save('超过最小计重费的30%.xlsx')
        number += 1
    print("已保存")


if __name__ == '__main__':
    data = []
    it = load_table("jifei.xlsx", "去除")
    sot = billing_weight_sort(it)
    for i in sot:
        # print(i)
        data.append(i)
    exceed_table(data)


